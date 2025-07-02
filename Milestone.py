import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from openpyxl.utils import column_index_from_string
from dateutil.relativedelta import relativedelta
from datetime import date
from milestone.veridia import *
from milestone.EwsLig import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from milestone.Eligo import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

COS_API_KEY = "ehl6KMyT95fwzKf7sPW_X3eKFppy_24xbm4P1Yk-jqyU"
COS_SERVICE_INSTANCE_ID = "crn:v1:bluemix:public:cloud-object-storage:global:a/fddc2a92db904306b413ed706665c2ff:e99c3906-0103-4257-bcba-e455e7ced9b7:bucket:projectreportnew"
COS_ENDPOINT = "https://s3.us-south.cloud-object-storage.appdomain.cloud"
COS_BUCKET = "projectreportnew"

cos_client = ibm_boto3.client(
    's3',
    ibm_api_key_id=COS_API_KEY,
    ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
    config=Config(signature_version='oauth'),
    endpoint_url=COS_ENDPOINT
)

def to_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
    return output.getvalue()

# Create Excel file content



foundewslig = False
foundeiden = False
foundeligo = False
foundverdia = False
foundeligog = False
foundeligoh = False
foundveridiaf4 = False
foundveridiaf5 = False


veridia = None
ews_lig = None
eligo = None

def get_cos_files():
    try:
        response = cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            print("No .json files found in the bucket 'ozonetell'. Please ensure JSON files are uploaded.")
        return files
    except Exception as e:
        print(f"Error fetching COS files: {e}")
        return ["Error fetching COS files"]
    

files = get_cos_files()
st.write(files)

today = date.today()
prev_month = today - relativedelta(months=1)

foundverdia = False

month_year = today.strftime("%m-%Y")
prev_month_year = prev_month.strftime("%m-%Y")


# Streamlit File Uploader
# uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

# if uploaded_file is not None:
#     # ProcessMilestone1(uploaded_file)
#     # ProcessEWSLIG(uploaded_file)
#     ProcessGandH(uploaded_file)


#=============VERIDIA================
for file in files:
                    
        try:
            if file.startswith("Veridia") and "Structure Work Tracker" in file and month_year in file:
                
                st.write("✅ Current month:", file)
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                veridia = ProcessMilestone1(io.BytesIO(response['Body'].read()))
                # st.write(veridia)
                foundverdia = True
                break
                    # elif prev_month_year in file:
                    #     st.write("🕓 Previous month:", file)
                    #     response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    #     veridia = ProcessVeridia(io.BytesIO(response['Body'].read()))
                    #     st.write(veridia)
        except Exception as e:
            st.info(e)

if not foundverdia:
    for file in files:
        try:
            if file.startswith("Veridia") and "Structure Work Tracker" in file and prev_month_year in file:
                st.write("🕓 Previous month:", file)
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                veridia = ProcessMilestone1(io.BytesIO(response['Body'].read()))
                # st.write(veridia)
                break
           
        except Exception as e:
            st.error(e)

#=============VERIDIA================


#===========EWS LIG=================

for file in files:
        try:
            if file.startswith("EWS LIG") and "Structure Work Tracker" in file and month_year in file:
                st.write("✅ Current month:", file)
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                ews_lig = ProcessEWSLIG(io.BytesIO(response['Body'].read()))
                # st.write(ews_lig)
                foundewslig = True
                break
        except Exception as e:
            st.error(e)

if not foundewslig:
    for file in files:
        try:
            if file.startswith("EWS LIG") and "Structure Work Tracker" in file and prev_month_year in file:
                st.write("🕓 Previous month:", file)
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                ews_lig = ProcessEWSLIG(io.BytesIO(response['Body'].read()))
                # st.write(ews_lig)
                break
            
        except Exception as e:
            st.error(e)

#===========EWS LIG=================


#===========ELIGO============
for file in files:
    
        try:
            if file.startswith("Eligo") and "Structure Work Tracker" in file and month_year in file:
                
                st.write("✅ Current month:", file)
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                eligo = ProcessGandH(io.BytesIO(response['Body'].read()))
                # st.write(eligo)
                foundeligo = True
                break
                    # elif prev_month_year in file:
                    #     st.write("🕓 Previous month:", file)
                    #     response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    #     eligo = ProcessGandH(io.BytesIO(response['Body'].read()))
                    #     st.write(eligo)
        except Exception as e:
            st.info(e)


if not foundeligo:
    for file in files:
        try:
            if file.startswith("Eligo") and "Structure Work Tracker" in file and prev_month_year in file:
                st.write("🕓 Previous month:", file)
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                eligo = ProcessGandH(io.BytesIO(response['Body'].read()))
                # st.write(eligo)
                break
        except Exception as e:
            st.error(e)
#===========ELIGO============

combined_json = []
missing_sources = []

st.write(combined_json)
if veridia is None:
    missing_sources.append("Veridia")
else:
    combined_json += veridia

if ews_lig is None:
    missing_sources.append("EWS_LIG")
else:
    combined_json += ews_lig

if eligo is None:
    missing_sources.append("Eligo")
else:
    combined_json += eligo

if missing_sources:
    st.warning(f"Missing Tow Files from: {', '.join(missing_sources)}")
else:
    st.success("All Datas loaded successfully.")

if combined_json:
    df = pd.DataFrame(combined_json)
    st.write(df)
    df['date'] = pd.to_datetime(df['date'], errors='coerce')


    df['year'] = df['date'].dt.year
    df['month'] = df['date'].dt.month

    month_map = {
        1: 'JAN', 2: 'FEB', 3: 'MAR', 4: 'APR',
        5: 'MAY', 6: 'JUN', 7: 'JUL', 8: 'AUG',
        9: 'SEP', 10: 'OCT', 11: 'NOV', 12: 'DEC'
    }

    df['month_name'] = df['month'].map(month_map)

    reverse_month_map = {v: k for k, v in month_map.items()}


    selected_year = st.sidebar.selectbox("Select Year", sorted(df['year'].unique()))


    selected_month_names = st.sidebar.multiselect(
        "Select Month(s)", 
        options=list(month_map.values()), 
        default=list(month_map.values()) 
    )


    selected_months = [reverse_month_map[m] for m in selected_month_names]

    filtered_df = df[(df['year'] == selected_year) & (df['month'].isin(selected_months))]

    # st.write(filtered_df)

    json_data = filtered_df.to_json(orient='records')
    # st.write(json_data)

    def process_json_data(json_data):
        # """
        # Process JSON data to create a structured format similar to the table
        # """
        # # Parse JSON if it's a string
        # if isinstance(json_data, str):
        #     data = json.loads(json_data)
        # else:
        #     data = json_data
        
        # Convert data to DataFrame
        df = pd.DataFrame(json_data)
        # st.write(df)
        
        # Handle different date formats
        if df['date'].dtype == 'object' and isinstance(df['date'].iloc[0], str):
            # If date is a string (like "Timestamp('2025-07-19 00:00:00')")
            df['date_clean'] = df['date'].str.extract(r"Timestamp\('([^']+)'\)")
            df['date_clean'] = pd.to_datetime(df['date_clean'])
        else:
            # If date is already datetime or timestamp
            df['date_clean'] = pd.to_datetime(df['date'])
        
        # Use original Tower names as modules
        df['module'] = df['Tower']
        # st.write(df['date_clean'])
        # Get unique months from the data and sort them
        unique_months = sorted(df['date_clean'].dt.month.unique())
        # st.write(unique_months)
        month_names = []
        month_mapping = {
            1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
            7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'
        }
        
        for month_num in unique_months:
            month_names.append(month_mapping[month_num])
        
        # Group by module and create monthly schedule
        result_data = []
        modules = df['module'].unique()
        
        for module in sorted(modules):
            module_data = df[df['module'] == module]
            
            # Initialize row
            row = {'Modules': module}
            
            # Check each month that exists in data
            for i, month_num in enumerate(unique_months):
                month_name = month_names[i]
                
                # Find entries for this month
                month_entries = module_data[module_data['date_clean'].dt.month == month_num]
                
                if len(month_entries) == 0:
                    row[month_name] = f'No work plan for {month_name.lower()[:3]}'
                else:
                    # Get floor information
                    floors = month_entries['floor'].unique()
                    row[month_name] = ', '.join(sorted(floors))
            
            result_data.append(row)
        
        return pd.DataFrame(result_data), month_names

    def create_excel_file(df, month_names):
        """
        Create Excel file with formatting similar to the image
        """
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        
        # Define styles
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        date_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        date_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add today's date in first row
        today_date = datetime.now().strftime("%B %d, %Y")
        date_cell = ws.cell(row=1, column=1, value=f"Downloaded on: {today_date}")
        date_cell.fill = date_fill
        date_cell.font = date_font
        date_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Merge cells for date across all columns
        headers = ['Modules'] + month_names
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        
        # Add headers dynamically in row 2
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Add data (starting from row 3 now)
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 3):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add totals row
        total_row = len(df) + 3
        ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
        
        # Calculate totals for each month
        for col in range(2, len(headers) + 1):  # Dynamic month columns
            month_data = [ws.cell(row=r, column=col).value for r in range(3, total_row)]
            # Count non-"No work plan" entries
            count = sum(1 for item in month_data if item and not item.startswith('No work plan'))
            ws.cell(row=total_row, column=col, value=f"{count} Slabs").font = Font(bold=True)
        
        # Auto-adjust column widths - FIXED VERSION
        for col_num in range(1, len(headers) + 1):
            column_letter = ws.cell(row=2, column=col_num).column_letter
            max_length = 0
            
            # Check all cells in this column (starting from row 2, skipping merged date row)
            for row_num in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb

    processed_df, month_names = process_json_data(filtered_df)

    st.subheader("Preview of processed data:")
    st.dataframe(processed_df)

    # Show detected months
    st.info(f"Detected months in your data: {', '.join(month_names)}")

    # Create Excel file
    wb = create_excel_file(processed_df, month_names)

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Download button
    st.download_button(
        label="📥 Download Excel File",
        data=excel_buffer.getvalue(),
        file_name=f"construction_schedule_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.error("Current or Previous Month Files not found for Veridia Structure / EWS_LIG Structure / Eligo Structure")

