
import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime,date
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config


# Function to process the Excel file - improved with consistent bold detection
def process_file_Tower5(file_stream):
    workbook = pd.ExcelFile(file_stream)
    
    if "TOWER 5 FINISHING." in workbook.sheet_names:
        sheet_name = "TOWER 5 FINISHING."
        
        # Read the data with header at row 1 (Excel row 1, 0-based index 0)
        df = pd.read_excel(file_stream, sheet_name=sheet_name, header=0)
        
        st.session_state.date = df[['Activity Name', 'Start', 'Finish']].head().iloc[1:2]
        # Assign column names based on document structure
        expected_columns = [
            'Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
            'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
            'Actual Start', 'Actual Finish', '% Complete', 'Start', 'Finish', 'Delay Reasons'
        ]
        
        if len(df.columns) >= len(expected_columns):
            df.columns = expected_columns[:len(df.columns)]
        else:
            st.error("Excel file has fewer columns than expected.")
            return None, None
        
        # Select desired columns
        target_columns = ["Module", "Floor", "Flat", "Activity ID", "Activity Name", "Start", "Finish"]
        df = df[target_columns]
        
        # List of activity names to filter
        activity_names = [
            "Brickwork",
            "AC Installation",
            "Balconies Waterproofing",
            "Brick masonry for entrance wall",
            "C-F-First Fix",
            "C-Gypsum and POP Punning",
            "C-P-First Fix",
            "C-Stone flooring",
            "Closing of shafts",
            "Copper Piping",
            "Counter stone works",
            "CP-Final Fix",
            "EL-Final Fix",
            "EL-Second Fix",
            "False ceiling framing",
            "Fixing of brackets for GRC Moduling",
            "Floor Tiling",
            "Glass Installation",
            "GRC jali fixing (Fire escape staicase)",
            "GRC jali fixing (main staircase)",
            "GRC jali fixing (splash pool)",
            "GRC molding fixing",
            "Grouting of toilets & balcony Tiles",
            "Gypsum board false ceiling",
            "Installation of Rear & Front balcony UPVC Windows",
            "Installation of doors",
            "Installation of wardrobes and cabinets",
            "Ledge Wall Construction",
            "MS works in balconies",
            "Paint in balcony and shafts",
            "Painting First Coat",
            "SS Framing",
            "ST-Electrical",
            "ST-Fire fighting",
            "ST-Plumbing & Water supply",
            "Stone cills, ledges and jambs",
            "Texture paint (final coat)",
            "Texture paint (first coat)",
            "Wall Tiling",
            "Water Proofing Works",
            "Waterproofing works"
        ]
        
        # Filter rows where Activity Name matches one of the specified activity_names
        df_filtered = df[df['Activity Name'].isin(activity_names)]
        
        # Convert 'Finish' column to datetime
        df_filtered['Finish'] = pd.to_datetime(df_filtered['Finish'], errors='coerce')
        
        # Extract month and year from the 'Finish' column
        df_filtered['Finish Month'] = df_filtered['Finish'].dt.strftime('%b')
        df_filtered['Finish Year'] = df_filtered['Finish'].dt.year
        
        return df_filtered[['Activity ID','Activity Name','Finish', 'Finish Month', 'Finish Year']], "Tower 5"
    
    else:
        return None, None