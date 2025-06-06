import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io


ews1 = []
ews2 = []
ews3 = []
lig1 = []
lig2 = []
lig3 = []

WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-3-70b-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KS5iR_XHOYc4N_xoId6YcXFjZR2ikINRdAyc2w2o18Oo"


def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    
    if response.status_code != 200:
        st.write(f"Failed to get access token: {response.text}")
        return None
    else:
        token_info = response.json()
        return token_info['access_token']


def generatePrompt(json_datas):
    body = {
        "input": f"""
         
        Read all data from this table carefully:
         
        {json_datas}.
        
        need a average value as percentage for green as single json  take poject name of each tower on that table
        
        Calculate the average value for green as a percentage and return the result in JSON format. Do not change the "Project" field value.

        For the "Structure" percentage, divide the green value by the non-green value.

        Use this formula:
        Structure = (Total Green / Total Non-Green) × 100

        Sample json:
       [{{
        ""
           "Project":"Project name"
           "Tower Name:"tower name",
           "Structure":"percentage %",
           "Finishing":"0%"
        }}]

        Return the result strictly as a JSON object—no code, no explanations, only the JSON.

        Dont put <|eom_id|> or any other

        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    
    if not headers["Authorization"]:
        return "Error: No valid access token."
    
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
    if response.status_code != 200:
        st.write(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    # st.write(json_datas)
    return response.json()['results'][0]['generated_text'].strip()

# def Checkcolor(sheet, )

def EWS1(sheet):
    st.write("Analyzing Ews Tower 1")
    rows = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]
    cols = ['D', 'H', 'L', 'P']
    flats = ["Pour 1", "Pour 2", "Pour 3", "Pour 4"]
    floors = ["GF", "1F", "2F", "3F", "4F", "5F", "6F", "7F", "8F", "9F", "10F", "11F", "12F", "13F", "14F"]
    for col in cols:
        
        for row in rows:
            cell = sheet[f"{col}{row}"]
            value = cell.value if cell.value is not None else ""
        
            fill = cell.fill
            if fill.start_color.type == 'rgb' and fill.start_color.rgb:
                bg_color = f"#{fill.start_color.rgb[-6:]}"  # Extract the last 6 chars for hex color
            else:
                bg_color = "#FFFFFF"  # Default to white if no background color is set
    
            # st.write(f"value:{value} --> color:{bg_color}")
            if bg_color == "#0070C0":
                ews1.append({
                    "date":value,
                    "color":bg_color,
                    "floor":floors[row-8],
                    "flat":flats[cols.index(col)-1],
                    "Tower":"EWS Tower 3"
                })

    

def EWS2(sheet):
    st.write("Analyzing Ews Tower 2")
    rows = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]
    cols = ['U','Y', 'AC', 'AG']
    flats = ["Pour 1", "Pour 2", "Pour 3", "Pour 4"]
    floors = ["GF", "1F", "2F", "3F", "4F", "5F", "6F", "7F", "8F", "9F", "10F", "11F", "12F", "13F", "14F"]
    for col in cols:
        
        for row in rows:
            cell = sheet[f"{col}{row}"]
            value = cell.value if cell.value is not None else ""
        
            fill = cell.fill
            if fill.start_color.type == 'rgb' and fill.start_color.rgb:
                bg_color = f"#{fill.start_color.rgb[-6:]}"  # Extract the last 6 chars for hex color
            else:
                bg_color = "#FFFFFF"  # Default to white if no background color is set
    
            if bg_color == "#0070C0":
                ews2.append({
                    "date":value,
                    "color":bg_color,
                    "floor":floors[row-8],
                    "flat":flats[cols.index(col)-1],
                    "Tower":"EWS Tower 3"
                })
    
    # st.write(len(ews2))
   

def EWS3(sheet):
    st.write("Analyzing Ews Tower 3")
    rows = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]
    cols = ['AL', 'AP', 'AT', 'AX']
    flats = ["Pour 1", "Pour 2", "Pour 3", "Pour 4"]
    floors = ["GF", "1F", "2F", "3F", "4F", "5F", "6F", "7F", "8F", "9F", "10F", "11F", "12F", "13F", "14F"]
    for col in cols:
        
        for row in rows:
            cell = sheet[f"{col}{row}"]
            value = cell.value if cell.value is not None else ""
        
            fill = cell.fill
            if fill.start_color.type == 'rgb' and fill.start_color.rgb:
                bg_color = f"#{fill.start_color.rgb[-6:]}"  # Extract the last 6 chars for hex color
            else:
                bg_color = "#FFFFFF"  # Default to white if no background color is set
    
            if bg_color == "#0070C0":
                ews3.append({
                    "date":value,
                    "color":bg_color,
                    "floor":floors[row-8],
                    "flat":flats[cols.index(col)-1],
                    "Tower":"EWS Tower 3"
                })
   

def LIG1(sheet):
    st.write("Analyzing Lig Tower 1")
    rows = [30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44]
    cols = ['AL', 'AP', 'AT', 'AX']
    flats = ["Pour 1", "Pour 2", "Pour 3", "Pour 4"]
    floors = ["GF", "1F", "2F", "3F", "4F", "5F", "6F", "7F", "8F", "9F", "10F", "11F", "12F", "13F", "14F"]
    for col in cols:
        
        for row in rows:
            cell = sheet[f"{col}{row}"]
            value = cell.value if cell.value is not None else ""
        
            fill = cell.fill
            if fill.start_color.type == 'rgb' and fill.start_color.rgb:
                bg_color = f"#{fill.start_color.rgb[-6:]}"  # Extract the last 6 chars for hex color
            else:
                bg_color = "#FFFFFF"  # Default to white if no background color is set
           
            if bg_color == "#0070C0":
                lig1.append({
                    "date":value,
                    "color":bg_color,
                    "floor":floors[row-30],
                    "flat":flats[cols.index(col)-1],
                    "Tower":"LIG TOWER 1"
                })
  
            
         
    

def LIG2(sheet):
    st.write("Analyzing Lig Tower 2")
    rows = [30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44]
    cols = ['U','Y', 'AC','AG']
    flats = ["Pour 1", "Pour 2", "Pour 3", "Pour 4"]
    floors = ["GF", "1F", "2F", "3F", "4F", "5F", "6F", "7F", "8F", "9F", "10F", "11F", "12F", "13F", "14F"]
    for col in cols:
        
        for row in rows:
            cell = sheet[f"{col}{row}"]
            value = cell.value if cell.value is not None else ""
        
            fill = cell.fill
            if fill.start_color.type == 'rgb' and fill.start_color.rgb:
                bg_color = f"#{fill.start_color.rgb[-6:]}"  # Extract the last 6 chars for hex color
            else:
                bg_color = "#FFFFFF"  # Default to white if no background color is set

            if bg_color == "#0070C0":
                lig2.append({
                    "date":value,
                    "color":bg_color,
                    "floor":floors[row-30],
                    "flat":flats[cols.index(col)-1],
                    "Tower":"LIG TOWER 2"
                })
    
            

def LIG3(sheet):
    st.write("Analyzing Lig Tower 3")
    rows = [30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44]
    cols = ['D', 'H', 'L', 'P']
    flats = ["Pour 1", "Pour 2", "Pour 3", "Pour 4"]
    floors = ["GF", "1F", "2F", "3F", "4F", "5F", "6F", "7F", "8F", "9F", "10F", "11F", "12F", "13F", "14F"]
    for col in cols:
        
        for row in rows:
            cell = sheet[f"{col}{row}"]
            value = cell.value if cell.value is not None else ""
        
            fill = cell.fill
            if fill.start_color.type == 'rgb' and fill.start_color.rgb:
                bg_color = f"#{fill.start_color.rgb[-6:]}"  # Extract the last 6 chars for hex color
            else:
                bg_color = "#FFFFFF"  # Default to white if no background color is set


            if bg_color == "#0070C0":
                lig3.append({
                    "date":value,
                    "color":bg_color,
                    "floor":floors[row-30],
                    "flat":flats[cols.index(col)-1],
                    "Tower":"LIG TOWER 3"
                })
    

def ProcessEWSLIG(exceldatas):

    wb = load_workbook(exceldatas, data_only=True)
    sheet_names = wb.sheetnames
    sheet_name = "Revised Baseline 45daysNGT+Rai"

    sheet = wb[sheet_name]

   
    EWS1(sheet)
    EWS2(sheet)
    EWS3(sheet)
    LIG1(sheet)
    LIG2(sheet)
    LIG3(sheet)

    all_data = ews1 + ews2 + ews3 + lig1 + lig2 + lig3
    
    # st.write(json_data)
    return all_data

    # st.write(ews1.count(1))

           

       

   

    


    



