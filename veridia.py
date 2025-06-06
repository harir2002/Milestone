import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from openpyxl.utils import column_index_from_string
from dateutil.relativedelta import relativedelta
from datetime import date
import xlwings as xw
import calendar

tower2 = []
tower3 = []
tower4 = []
tower5 = []
tower6 = []
tower7 = []


WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-3-70b-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KS5iR_XHOYc4N_xoId6YcXFjZR2ikINRdAyc2w2o18Oo"


def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    
    if response.status_code != 200:
        st.write(f"Failed to get access token: {response.text}")
        return None
    else:
        token_info = response.json()
        return token_info['access_token']


def generatePrompt(json_datas):
    body = {
        "input": f"""
         
        Read all data from this table carefully:
         
        {json_datas}.
        
        need a average value as percentage for green as single json  take poject name of each tower on that table
        
        Calculate the average value for green as a percentage and return the result in JSON format. Do not change the "Project" field value.

        For the "Structure" percentage, divide the green value by the non-green value.

        Use this formula:
        Structure = (Total Green / Total Non-Green) × 100

        Sample json:
       [[{{
        ""
           "Project":"Project name"
           "Tower Name:"tower name",
           "Structure":"percentage %",
           "Finishing":"0%"
        }}]

        Return the result strictly as a JSON object—no code, no explanations, only the JSON.

        Dont put <|eom_id|> or any other

        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    
    if not headers["Authorization"]:
        return "Error: No valid access token."
    
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
    if response.status_code != 200:
        st.write(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    # st.write(json_datas)
    return response.json()['results'][0]['generated_text'].strip()



#FF00B0F0 - Blue


def Tower2(sheet):
    st.write("Analyzing Veridia Tower 2")
    rows = [3, 4, 5, 6, 7, 9, 10, 12, 14, 15, 16, 17, 19, 20]
    cols = ['B', 'D', 'F', 'H', 'J', 'L', 'N', 'P']
    current_flat = ""  # Initialize current_flat

    # Map rows to floors (cycle 1 to 6)
    floor_mapping = {row: (i % 6) + 1 for i, row in enumerate(rows)}  # Cycles floors 1 to 6

    for col in cols:
        for row in rows:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            cell_value = cell.value if cell.value is not None else ""  # Handle None values

           
            if fill.fill_type == "solid" and fill.start_color and fill.start_color.rgb:
                color = fill.start_color.rgb  # Get RGB color code
            else:
                color = "no solid fill"

            if isinstance(cell_value, str) and (cell_value.lower().endswith("east") or cell_value.lower().endswith("west") or cell_value.lower().endswith("north") or cell_value.lower().endswith("south") or cell_value.lower().startswith("east") or cell_value.lower().startswith("west") or cell_value.lower().startswith("north") or cell_value.lower().startswith("south")):
                
                current_flat = cell_value

            floor = floor_mapping[row]

            tower2.append({
                "date": cell_value,
                "color": color,
                "floor": f"{floor}F",
                "flat": current_flat,
                "Tower":"Veridia Tower2"
            })

        

    # st.write(f"Total entries: {len(tower2)}")
    # st.dataframe(tower2)
    return tower2
   


def Tower3(sheet):
    st.write("Analyzing Veridia Tower 3")
    rows = [3, 4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['T', 'V', 'X', 'Z', 'AB', 'AD', 'AF', 'AH']
    current_flat = ""  # Initialize current_flat

    floor_mapping = {row: (i % 6) + 1 for i, row in enumerate(rows)}  # Cycles floors 1 to 6

    for col in cols:
        for row in rows:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            cell_value = cell.value if cell.value is not None else ""  # Handle None values

           
            if fill.fill_type == "solid" and fill.start_color and fill.start_color.rgb:
                color = fill.start_color.rgb  # Get RGB color code
            else:
                color = "no solid fill"

            if isinstance(cell_value, str) and (cell_value.lower().endswith("east") or cell_value.lower().endswith("west") or cell_value.lower().endswith("north") or cell_value.lower().endswith("south") or cell_value.lower().startswith("east") or cell_value.lower().startswith("west") or cell_value.lower().startswith("north") or cell_value.lower().startswith("south")):
                current_flat = cell_value

            floor = floor_mapping[row]

            tower3.append({
               "date": cell_value,
                "color": color,
                "floor": f"{floor}F",
                "flat": current_flat,
                "Tower":"Veridia Tower3"
            })

        

    # st.write(f"Total entries: {len(tower3)}")
    # st.dataframe(tower3)
    return tower3
            
    # for row in rows:

    #     for col in cols:
    #         col_idx = column_index_from_string(col)
    #         cell = sheet.cell(row=row, column=col_idx)
    #         value = cell.value
    #         color = get_fill_color(cell)
    #         # st.write(f"{col}{row}")
    #         # st.write(value)
    #         # st.write(color)
    #         if color == "Theme (9)" or color == "Theme (7)" or color == "Theme (5)" or color == "FF92D050" or color == "FFFFFF00":
    #             tower3.append(1)
    #         else:
    #             tower3.append(0)

def Tower4(sheet):
    st.write("Analyzing Veridia Tower 4")
    rows = [3, 4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['AL', 'AN', 'AP', 'AR', 'AT', 'AV', 'AX', 'AZ', 'BB', 'BD', 'BF', 'BH', 'BJ', 'BL', 'BN', 'BP']
    
    current_flat = ""  # Initialize current_flat
    floor_mapping = {row: (i % 6) + 1 for i, row in enumerate(rows)}  # Cycles floors 1 to 6

    for col in cols:
        for row in rows:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            cell_value = cell.value if cell.value is not None else ""  # Handle None values

           
            if fill.fill_type == "solid" and fill.start_color and fill.start_color.rgb:
                color = fill.start_color.rgb  # Get RGB color code
            else:
                color = "no solid fill"

            if isinstance(cell_value, str) and (cell_value.lower().endswith("east") or cell_value.lower().endswith("west") or cell_value.lower().endswith("north") or cell_value.lower().endswith("south") or cell_value.lower().startswith("east") or cell_value.lower().startswith("west") or cell_value.lower().startswith("north") or cell_value.lower().startswith("south")):
                # st.write(cell_value.lower())
                current_flat = cell_value

            floor = floor_mapping[row]

            tower4.append({
                "date": cell_value,
                "color": color,
                "floor": f"{floor}F",
                "flat": current_flat,
                "Tower":"Veridia Tower4"
            })

        

    # st.write(f"Total entries: {len(tower4)}")
    # st.dataframe(tower4)
    return tower4
                

def Tower5(sheet):
    st.write("Analyzing Veridia Tower 5")
    rows = [3, 4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['DC', 'DE', 'DG', 'DI', 'DK', 'DM', 'DO', 'DQ', 'DS', 'DU', 'DW', 'DY', 'EA', 'EC']
    current_flat = ""  # Initialize current_flat

    floor_mapping = {row: (i % 6) + 1 for i, row in enumerate(rows)}  # Cycles floors 1 to 6

    for col in cols:
        for row in rows:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            cell_value = cell.value if cell.value is not None else ""  # Handle None values

           
            if fill.fill_type == "solid" and fill.start_color and fill.start_color.rgb:
                color = fill.start_color.rgb  # Get RGB color code
            else:
                color = "no solid fill"

            if isinstance(cell_value, str) and (cell_value.lower().endswith("east") or cell_value.lower().endswith("west") or cell_value.lower().endswith("north") or cell_value.lower().endswith("south") or cell_value.lower().startswith("east") or cell_value.lower().startswith("west") or cell_value.lower().startswith("north") or cell_value.lower().startswith("south")):
                current_flat = cell_value

            floor = floor_mapping[row]

            tower5.append({
                "date": cell_value,
                "color": color,
                "floor": f"{floor}F",
                "flat": current_flat,
                "Tower":"Veridia Tower5"
            })

        

    # st.write(f"Total entries: {len(tower5)}")
    # st.dataframe(tower5)
    return tower5
                

def Tower6(sheet):
    st.write("Analyzing Veridia Tower 6")
    rows = [3, 4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['FK', 'FM', 'FO', 'FQ', 'FS', 'FU', 'FW', 'FY',  'GA', 'GC', 'GE', 'GG', 'GI', "GK"]
    current_flat = ""  # Initialize current_flat

    floor_mapping = {row: (i % 6) + 1 for i, row in enumerate(rows)}  # Cycles floors 1 to 6

    for col in cols:
        for row in rows:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            cell_value = cell.value if cell.value is not None else ""  # Handle None values

           
            if fill.fill_type == "solid" and fill.start_color and fill.start_color.rgb:
                color = fill.start_color.rgb  # Get RGB color code
            else:
                color = "no solid fill"

            if isinstance(cell_value, str) and (cell_value.lower().endswith("east") or cell_value.lower().endswith("west") or cell_value.lower().endswith("north") or cell_value.lower().endswith("south") or cell_value.lower().startswith("east") or cell_value.lower().startswith("west") or cell_value.lower().startswith("north") or cell_value.lower().startswith("south")):
                current_flat = cell_value

            floor = floor_mapping[row]

            tower6.append({
                "date": cell_value,
                "color": color,
                "floor": f"{floor}F",
                "flat": current_flat,
                "Tower":"Veridia Tower6"
            })

        

    # st.write(f"Total entries: {len(tower6)}")
    # st.dataframe(tower6)
    return tower6
                

def Tower7(sheet):
    st.write("Analyzing Veridia Tower 7")

    rows = [3, 4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['EG', 'EI', 'EK', 'EM', 'EO', 'EQ', 'ES', 'EU', 'EW', 'EY', 'FA', 'FC', 'FE', 'FG']
    current_flat = ""  # Initialize current_flat
    
    floor_mapping = {row: (i % 6) + 1 for i, row in enumerate(rows)}  # Cycles floors 1 to 6

    for col in cols:
        for row in rows:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            cell_value = cell.value if cell.value is not None else ""  # Handle None values

           
            if fill.fill_type == "solid" and fill.start_color and fill.start_color.rgb:
                color = fill.start_color.rgb  # Get RGB color code
            else:
                color = "no solid fill"

            if isinstance(cell_value, str) and (cell_value.lower().endswith("east") or cell_value.lower().endswith("west") or cell_value.lower().endswith("north") or cell_value.lower().endswith("south") or cell_value.lower().startswith("east") or cell_value.lower().startswith("west") or cell_value.lower().startswith("north") or cell_value.lower().startswith("south")):
                current_flat = cell_value

            floor = floor_mapping[row]

            tower7.append({
                "date": cell_value,
                "color": color,
                "floor": f"{floor}F",
                "flat": current_flat,
                "Tower":"Veridia Tower7"
            })

        

    # st.write(f"Total entries: {len(tower7)}")
    # st.dataframe(tower7)
    return tower7
                

def ProcessMilestone1(exceldatas):
    wb = load_workbook(exceldatas, data_only=True)
    sheet_names = wb.sheetnames
    sheet_name = "Revised baseline with 60d NGT"

    sheet = wb[sheet_name]

    Tower2(sheet)
    st.divider()
    Tower3(sheet)
    st.divider()
    Tower4(sheet)
    st.divider()
    Tower5(sheet)
    st.divider()
    # st.write(test)
    Tower6(sheet)
    st.divider()
    Tower7(sheet)

    all_datas = tower2 + tower3 + tower4 + tower5 + tower6 + tower7
    
    
    return all_datas

