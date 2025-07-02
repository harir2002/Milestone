import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from openpyxl.utils import column_index_from_string
from dateutil.relativedelta import relativedelta
from datetime import date
from veridia import *
from EwsLig import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from Eligo import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from Tower4 import *
from Tower5 import *
from Tower4 import process_file_Tower4
from Tower5 import process_file_Tower5
import re

def process_activity_data(data):
    """Process JSON data and create activity count by month"""
    df = pd.DataFrame(data)
    
    # Convert finish date to datetime
    df['Finish_Date'] = pd.to_datetime(df['Finish'])
    
    # Create a pivot table to count activities by name and month
    pivot_df = df.groupby(['Activity Name', 'Finish Month']).size().reset_index(name='Count')
    
    # Create the final dataframe in the desired format
    activity_summary = pivot_df.pivot(index='Activity Name', columns='Finish Month', values='Count').fillna(0)
    
    # Convert to integers
    activity_summary = activity_summary.astype(int)
    
    # Reset index to make Activity Name a column
    activity_summary = activity_summary.reset_index()
    
    return activity_summary

def create_tower_format(activity_summary, tower_name="Tower-ABC"):
    """Create the Tower format similar to your image"""
    # Create milestone rows
    result_data = []
    
    # Create proper header row - Tower name and month columns in chronological order
    # Define month order
    month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    # Get month columns from original data and sort them chronologically
    original_months = list(activity_summary.columns[1:])
    sorted_months = [month for month in month_order if month in original_months]
    
    tower_row = [tower_name, "Activity Name"] + sorted_months
    result_data.append(tower_row)
    
    # Add each activity as a milestone
    for idx, row in activity_summary.iterrows():
        milestone_name = f"Milestone-{idx + 1}"
        activity_name = row['Activity Name']
        # Reorder the counts to match the sorted months
        counts = [f"{int(row[month])} activities" if row[month] > 0 else "No activities" 
                 for month in sorted_months]
        
        milestone_row = [milestone_name, activity_name] + counts
        result_data.append(milestone_row)
    
    # Create DataFrame with proper structure
    final_df = pd.DataFrame(result_data[1:], columns=result_data[0])
    
    return final_df

def to_excel(df):
    """Convert DataFrame to Excel bytes"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Activity Summary')
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Activity Summary']
        
        # Style the header row (row 2 since we'll insert date in row 1)
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        
        # Insert a new row at the top for the date
        worksheet.insert_rows(1)
        
        # Add today's date in the first row, second column
        today_date = datetime.now().strftime("%B %d, %Y")
        date_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        date_font = Font(bold=True)
        
        # Place date in row 1, column 2 (B1)
        date_cell = worksheet.cell(row=1, column=2, value=f"Downloaded on: {today_date}")
        date_cell.fill = date_fill
        date_cell.font = date_font
        date_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Style the actual header row (now row 2)
        for cell in worksheet[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output.getvalue()

COS_API_KEY = "ehl6KMyT95fwzKf7sPW_X3eKFppy_24xbm4P1Yk-jqyU"
COS_SERVICE_INSTANCE_ID = "crn:v1:bluemix:public:cloud-object-storage:global:a/fddc2a92db904306b413ed706665c2ff:e99c3906-0103-4257-bcba-e455e7ced9b7:bucket:projectreportnew"
COS_ENDPOINT = "https://s3.us-south.cloud-object-storage.appdomain.cloud"
COS_BUCKET = "projectreportnew"


cos_client = ibm_boto3.client(
    's3',
    ibm_api_key_id=COS_API_KEY,
    ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
    config=Config(signature_version='oauth'),
    endpoint_url=COS_ENDPOINT
)

def get_cos_files():
    try:
        response = cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            print("No .json files found in the bucket 'ozonetell'. Please ensure JSON files are uploaded.")
        return files
    except Exception as e:
        print(f"Error fetching COS files: {e}")
        return ["Error fetching COS files"]
    

if 'tower5df' not in st.session_state:
    st.session_state.tower5df = pd.DataFrame()
if 'tower4df' not in st.session_state:
    st.session_state.tower4df = pd.DataFrame()

foundveridiaf4 = False
foundveridiaf5 = False

today = date.today()
prev_month = today - relativedelta(months=1)

foundverdia = False

month_year = today.strftime("%m-%Y")
prev_month_year = prev_month.strftime("%m-%Y")


files = get_cos_files()
st.write(files)

tower4df = None
tower5df = None

def extract_date_from_filename(filename):
    # Match date in format (dd-mm-yyyy) or (dd-mm-yyyy) with possible spaces
    match = re.search(r'\((\d{2})-(\d{2})-(\d{4})\)', filename.replace(" ", ""))
    if match:
        day, month, year = match.groups()
        try:
            return datetime(int(year), int(month), int(day))
        except ValueError:
            return None
    return None

# Find latest file for each tower
latest_tower4_file = None
latest_tower4_date = None
latest_tower5_file = None
latest_tower5_date = None

for file in files:
    if "Tower 4 Finishing Tracker" in file:
        file_date = extract_date_from_filename(file)
        if file_date and (latest_tower4_date is None or file_date > latest_tower4_date):
            latest_tower4_date = file_date
            latest_tower4_file = file
    elif "Tower 5 Finishing Tracker" in file:
        file_date = extract_date_from_filename(file)
        if file_date and (latest_tower5_date is None or file_date > latest_tower5_date):
            latest_tower5_date = file_date
            latest_tower5_file = file

# Now process only the latest files
if latest_tower5_file:
    st.write("✅ Latest Tower 5 file:", latest_tower5_file)
    response = cos_client.get_object(Bucket="projectreportnew", Key=latest_tower5_file)
    st.session_state.tower5df, name = process_file_Tower5(io.BytesIO(response['Body'].read()))

if latest_tower4_file:
    st.write("✅ Latest Tower 4 file:", latest_tower4_file)
    response = cos_client.get_object(Bucket="projectreportnew", Key=latest_tower4_file)
    st.session_state.tower4df, name = process_file_Tower4(io.BytesIO(response['Body'].read()))



# tower4json = tower4df.to_json(orient='records')
# tower5json = tower5df.to_json(orient='records')

# Combine available months and years from both dataframes
all_months = set()
all_years = set()

for key in ['tower5df', 'tower4df']:
    if key in st.session_state and not st.session_state[key].empty:
        df = st.session_state[key]
        all_months.update(df['Finish Month'].dropna().unique())
        all_years.update(df['Finish Year'].dropna().unique())

# Sort for display
all_months = sorted(all_months)
all_years = sorted(all_years)

# Sidebar filter UI
with st.sidebar:
    selected_months = st.multiselect("Select Finish Month(s)", all_months, default=all_months)
    selected_year = st.selectbox("Select Finish Year", all_years)

# Initialize empty DataFrames
filtered_df5 = pd.DataFrame()
filtered_df4 = pd.DataFrame()

# Filter and display tower5df
if 'tower5df' in st.session_state and not st.session_state.tower5df.empty:
    df5 = st.session_state.tower5df
    filtered_df5 = df5[
        (df5['Finish Month'].isin(selected_months)) &
        (df5['Finish Year'] == selected_year)
    ]
    st.subheader("Filtered tower5df")
    st.write(filtered_df5)

# Filter and display tower4df
if 'tower4df' in st.session_state and not st.session_state.tower4df.empty:
    df4 = st.session_state.tower4df
    filtered_df4 = df4[
        (df4['Finish Month'].isin(selected_months)) &
        (df4['Finish Year'] == selected_year)
    ]
    st.subheader("Filtered tower4df")
    st.write(filtered_df4)

# Combine and display as JSON
combined_df = pd.concat([filtered_df5, filtered_df4], ignore_index=True)

if not combined_df.empty:
    st.subheader("Combined Filtered Data (as JSON)")
    data = combined_df.to_dict(orient='records')
    activity_summary = process_activity_data(data)
    final_df = create_tower_format(activity_summary)
    
    # Display the formatted table
    st.subheader("Activity Summary by Month")
    st.dataframe(final_df, use_container_width=True)

    excel_data = to_excel(final_df)
    
    st.download_button(
        label="📥 Download as Excel",
        data=excel_data,
        file_name=f"activity_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("No data matching the selected filters.")
