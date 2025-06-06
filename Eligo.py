import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from openpyxl.utils import column_index_from_string

towerf = []
towerg = []
towerh = []


WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-3-70b-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KS5iR_XHOYc4N_xoId6YcXFjZR2ikINRdAyc2w2o18Oo"


def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    
    if response.status_code != 200:
        st.write(f"Failed to get access token: {response.text}")
        return None
    else:
        token_info = response.json()
        return token_info['access_token']


def generatePrompt(json_datas):
    body = {
        "input": f"""
         
        Read all data from this table carefully:
         
        {json_datas}.
        
        need a average value as percentage for green as single json  take poject name of each tower on that table
        
        Calculate the average value for green as a percentage and return the result in JSON format. Do not change the "Project" field value.

        For the "Structure" percentage, divide the green value by the non-green value.

        Use this formula:
        Structure = (Total Green / Total Non-Green) × 100

        Sample json:
       [[{{
        ""
           "Project":"Project name"
           "Tower Name:"tower name",
           "Structure":"percentage %",
           "Finishing":"0%"
        }}]

        Return the result strictly as a JSON object—no code, no explanations, only the JSON.

        Dont put <|eom_id|> or any other

        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    
    if not headers["Authorization"]:
        return "Error: No valid access token."
    
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
    if response.status_code != 200:
        st.write(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    # st.write(json_datas)
    return response.json()['results'][0]['generated_text'].strip()


#FF00B0F0 = blue

def TowerF(sheet):
    # st.write("Analyzing Eligo Tower F")
    rows = [5, 6, 7, 8, 9, 10, 11, 12]
    c = 0
    cols = ['D', 'H']
    flats = ["Pour 1", "Pour 2"]
    floors = ["1F","2F", "3F", "4F", "5F", "6F", "7F", "8F"]
    
    for col in cols:
        for row in rows:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb  # Get RGB color code
                # st.write(f"Cell {col}{row} color: {color}, value: {cell.value}")
                if color == "FF00B0F0":
                    towerf.append({
                        "date":cell.value,
                        "color":color,
                        "floor":floors[row-5],
                        "flat":flats[row-5],
                        "Tower":f"Tower H {flats[c]}"

                    })
        c = c + 1
              
    

def TowerG(sheet):
    # st.write("Analyzing Eligo Tower G")
    rows = [5, 6, 7, 8, 9, 10, 11, 12]
    cols = ['N', 'R', 'V']
    c = 0
    flats = ["Pour 1", "Pour 2", "Pour 3"]
    floors = ["1F","2F", "3F", "4F", "5F", "6F", "7F", "8F"]

    for col in cols:
        for row in rows:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb  # Get RGB color code
                if color == "FF00B0F0":
                    towerg.append({
                        "date":cell.value,
                        "color":color,
                        "floor":floors[row - 5],
                        "flat":flats[c],
                        "Tower":f"Tower G {flats[c]}"
                    })
        c = c + 1
              
                

def TowerH(sheet):
    # st.write("Analyzing Eligo Tower H")
    rows = [5, 6, 7, 8, 9, 10, 11, 12]
    cols = ['AB', 'AF', 'AJ', 'AN', 'AR', 'AV', 'AZ']
    c = 0
    flats = ["Pour 7", "Pour 6", "Pour 5", "Pour 4", "Pour 3", "Pour 2", "Pour 1"]
    floors = ["1F","2F", "3F", "4F", "5F", "6F", "7F", "8F"]

    for col in cols:
        for row in rows:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb  # Get RGB color code
                # st.write(f"Cell {col}{row} color: {color}, value: {cell.value}")
                if color == "FF00B0F0":
                    towerh.append({
                        "date":cell.value,
                        "color":color,
                        "floor":floors[row-5],
                        "flat":flats[c],
                        "Tower": f"Tower H {flats[c]}"

                    })
        c = c + 1
              
    
               
           

def ProcessGandH(exceldatas):

    wb = load_workbook(exceldatas, data_only=True)
    sheet_names = wb.sheetnames
    sheet_name = "Revised Baselines- 25 days SC"

    sheet = wb[sheet_name]
    #Revised Baselines- 25 days SC
    TowerF(sheet)
    TowerG(sheet)
    TowerH(sheet)

    all_datas = towerf + towerg + towerh

    # st.dataframe(all_datas)
  
    return all_datas
        