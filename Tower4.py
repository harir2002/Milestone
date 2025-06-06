
import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime,date
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config


def process_file_Tower4(file_stream):
    # First, use pandas to read the data
    excel_file = pd.ExcelFile(file_stream)
    
    if "TOWER 4 FINISHING." in excel_file.sheet_names:
        sheet_name = "TOWER 4 FINISHING."
        df = pd.read_excel(file_stream, sheet_name=sheet_name, header=0)
        
        # Save date for Streamlit session state (optional)
        st.session_state.date = df[['Activity Name', 'Start', 'Finish']].head().iloc[1:2]

        # Assign expected columns
        expected_columns = [
            'Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
            'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
            'Actual Start', 'Actual Finish', '% Complete', 'Start', 'Finish', 'Delay Reasons'
        ]
        if len(df.columns) >= len(expected_columns):
            df.columns = expected_columns[:len(df.columns)]
        else:
            st.error("Excel file has fewer columns than expected.")
            return None, None

        # Filter only necessary columns
        target_columns = ["Module", "Floor", "Flat", "Activity ID", "Activity Name", "Start", "Finish"]
        df = df[target_columns]
        
        # Rewind stream to read again with openpyxl
        file_stream.seek(0)
        workbook = load_workbook(file_stream, data_only=True)
        ws = workbook[sheet_name]
        
        # Index for 'Activity Name' (Excel column F = index 6 in openpyxl, since it's 1-based)
        activity_col_idx = 6

        non_bold_rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, max_col=16), start=0):
            cell = row[activity_col_idx - 1]  # Convert 1-based to 0-based
            if cell.value and (not cell.font or not cell.font.bold):
                non_bold_rows.append(i)

        # Extract non-bold rows
        if non_bold_rows:
            df_non_bold = df.iloc[non_bold_rows]
        else:
            df_non_bold = pd.DataFrame(columns=df.columns)

        # Convert and extract date parts
        df_non_bold['Finish'] = pd.to_datetime(df_non_bold['Finish'], errors='coerce')
        df_non_bold['Finish Month'] = df_non_bold['Finish'].dt.strftime('%b')
        df_non_bold['Finish Year'] = df_non_bold['Finish'].dt.year

        return df_non_bold[['Activity ID','Activity Name','Finish', 'Finish Month', 'Finish Year']], "Tower 4"
